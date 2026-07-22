"""
Readers for the source formats.

SOV score channels:
  parse_linkedin(path, config)  -> (rows_by_brand, period_start, period_end, extras_skipped)
  parse_semrush(paths, config)  -> (rows_by_brand, period_label, missing_brands)

Degreed-only trend tabs (NOT part of the SOV score):
  parse_seo(path, config)       -> {"months": [...], "latest_ym": str, "domain": str}
  parse_aeo(path)               -> {"months": [...], "soa_by_ym": {...}, "latest_ym": str}
  parse_hubspot_forms(paths, config) -> {"months": [...], "patterns": [...]}

The SOV parsers return data keyed by the internal brand key from
config["brand_name_map"]. None of these functions does any share math -- that
lives in calc.py.
"""
import csv
import os
import re


_MON_ABBR = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
             "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
_MON_FULL = {"january": 1, "february": 2, "march": 3, "april": 4, "may": 5,
             "june": 6, "july": 7, "august": 8, "september": 9, "october": 10,
             "november": 11, "december": 12}
_MON_NAME = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
             7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _to_int(value):
    """Cast a possibly-float / possibly-string cell to int. Blank -> None."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip().replace(",", "")
        if value == "":
            return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _clean_number(text):
    """'148,383' -> 148383.0 ; '' -> None."""
    if text is None:
        return None
    text = str(text).strip().replace(",", "").replace("%", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _duration_to_seconds(text):
    """'07:45' -> 465 ; 'mm:ss' or 'hh:mm:ss'. Returns None if unparseable."""
    if not text:
        return None
    parts = str(text).strip().split(":")
    try:
        parts = [int(p) for p in parts]
    except ValueError:
        return None
    seconds = 0
    for p in parts:
        seconds = seconds * 60 + p
    return seconds


def _lookup_maps(config, channel):
    """Return {source_label_lower: internal_brand_key} for a channel."""
    out = {}
    for brand, labels in config["brand_name_map"].items():
        label = labels.get(channel)
        if label:
            out[label.strip().lower()] = brand
    return out


# --------------------------------------------------------------------------
# LinkedIn
# --------------------------------------------------------------------------
def parse_linkedin(path, config):
    """
    Read the COMPETITORS sheet of a LinkedIn export.

    Row 1: period start / end in the first two cells.
    Row 2: header (Page, Total Followers, New Followers,
                   Total post engagements, Total posts).
    Following rows: one per brand page.

    Returns (rows_by_brand, period_start, period_end, extras_skipped)
      rows_by_brand[brand] = {
          "total_followers", "new_followers", "engagements", "total_posts"
      }
    """
    import warnings
    from openpyxl import load_workbook

    # NB: not read_only. Some LinkedIn exports store a wrong/blank sheet
    # dimension, and read_only mode trusts it and yields only the header row.
    # These files are tiny, so loading fully is cheap and correct.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # "no default style" cosmetic warning
        wb = load_workbook(path, data_only=True)
    if "COMPETITORS" in wb.sheetnames:
        ws = wb["COMPETITORS"]
    else:
        ws = wb.active  # be forgiving if the tab was renamed

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        raise ValueError("LinkedIn sheet has no data rows: %s" % path)

    def _fmt_date(cell):
        if cell is None:
            return None
        if hasattr(cell, "strftime"):
            return cell.strftime("%-m/%-d/%Y")
        return str(cell).strip()

    period_start = _fmt_date(rows[0][0]) if rows[0] else None
    period_end = _fmt_date(rows[0][1]) if rows[0] and len(rows[0]) > 1 else None

    # Locate columns by header name so extra/reordered columns don't break us.
    header = [str(c).strip().lower() if c is not None else "" for c in rows[1]]

    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None

    c_page = col("page")
    c_foll = col("total followers")
    c_new = col("new followers")
    c_eng = col("total post engagements", "total engagements")
    c_posts = col("total posts")
    if c_page is None or c_foll is None or c_new is None or c_eng is None:
        raise ValueError("LinkedIn header missing expected columns: %s" % header)

    label_map = _lookup_maps(config, "linkedin")
    rows_by_brand = {}
    extras_skipped = []

    for r in rows[2:]:
        if r is None or c_page >= len(r):
            continue
        page = r[c_page]
        if page is None or str(page).strip() == "":
            continue
        key = label_map.get(str(page).strip().lower())
        if key is None:
            extras_skipped.append(str(page).strip())  # not in config -> skip
            continue
        rows_by_brand[key] = {
            "total_followers": _to_int(r[c_foll]),
            "new_followers": _to_int(r[c_new]),
            "engagements": _to_int(r[c_eng]),
            "total_posts": _to_int(r[c_posts]) if c_posts is not None
            and c_posts < len(r) else None,
        }

    return rows_by_brand, period_start, period_end, extras_skipped


# --------------------------------------------------------------------------
# Semrush
# --------------------------------------------------------------------------
def parse_semrush(paths, config):
    """
    Read one or more Semrush CSVs, merge, de-duplicate on domain.

    Each brand occupies two rows: the current month (non-empty Target),
    then the prior month on a blank-Target row that belongs to the brand
    above it. We keep only the current-month row for scoring.

    Returns (rows_by_brand, period_label, missing_brands)
      rows_by_brand[brand] = {
          "domain", "period", "visits", "unique_visitors",
          "avg_visit_duration_sec", "avg_visit_duration_raw", "bounce_rate"
      }
      missing_brands = configured brands with a web domain but no row found.
    """
    label_map = _lookup_maps(config, "web")
    rows_by_brand = {}
    period_label = None

    for path in paths:
        if not os.path.exists(path):
            continue
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                continue
            hmap = {str(h).strip().lower(): i for i, h in enumerate(header)}

            def idx(*names):
                for n in names:
                    if n in hmap:
                        return hmap[n]
                return None

            i_target = idx("target")
            i_period = idx("period")
            i_visits = idx("visits")
            i_uv = idx("unique visitors")
            i_dur = idx("avg. visit duration", "avg visit duration")
            i_bounce = idx("bounce rate")
            if i_target is None or i_uv is None:
                raise ValueError("Semrush header missing Target/Unique "
                                 "Visitors: %s" % path)

            for row in reader:
                if not row or i_target >= len(row):
                    continue
                target = (row[i_target] or "").strip()
                if target == "":
                    continue  # prior-month row; skip (belongs to row above)
                key = label_map.get(target.lower())
                if key is None:
                    continue  # domain not in config -> skip
                if key in rows_by_brand:
                    continue  # already captured (e.g. degreed.com repeated)

                dur_raw = row[i_dur].strip() if i_dur is not None \
                    and i_dur < len(row) else None
                rows_by_brand[key] = {
                    "domain": target,
                    "period": (row[i_period].strip() if i_period is not None
                               and i_period < len(row) else None),
                    "visits": _clean_number(row[i_visits]) if i_visits is not None
                    and i_visits < len(row) else None,
                    "unique_visitors": _clean_number(row[i_uv]),
                    "avg_visit_duration_sec": _duration_to_seconds(dur_raw),
                    "avg_visit_duration_raw": dur_raw,
                    "bounce_rate": _clean_number(row[i_bounce]) if i_bounce
                    is not None and i_bounce < len(row) else None,
                }
                if period_label is None:
                    period_label = rows_by_brand[key]["period"]

    # Which configured brands have a web domain but produced no row?
    configured_web = {b for b, m in config["brand_name_map"].items()
                      if m.get("web")}
    missing_brands = sorted(configured_web - set(rows_by_brand.keys()))

    return rows_by_brand, period_label, missing_brands


# --------------------------------------------------------------------------
# SEO (Degreed-only tab) -- SEMrush export, section 13a
# --------------------------------------------------------------------------
def _month_label_to_ym(text):
    """'Jun 2026' / 'June 2026' / '2026-06' -> '2026-06'. None if unparseable."""
    if not text:
        return None
    text = str(text).strip()
    m = re.match(r"^(\d{4})-(\d{1,2})$", text)
    if m:
        return "%04d-%02d" % (int(m.group(1)), int(m.group(2)))
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", text)
    if m:
        name = m.group(1).lower()
        num = _MON_FULL.get(name) or _MON_ABBR.get(name[:3])
        if num:
            return "%04d-%02d" % (int(m.group(2)), num)
    return None


def parse_seo(path, config):
    """
    Read a SEMrush export and return Degreed's monthly organic-search trend.

    The configured folder currently holds a "unified domain data" export in
    LONG format: one row per (Month, Domain) with metric columns
        Month, Domain, Authority Score, Semrush Rank, Organic Traffic,
        Organic Keywords, Backlinks, Referring Domains, Paid Keywords,
        Paid Traffic Cost
    We filter to the Degreed domain, keep the organic + link metrics, and drop
    the paid columns (out of scope for this tab, per section 13a).

    Returns {"months": [ {ym, month_label, organic_traffic, organic_keywords,
                          backlinks, referring_domains, authority_score,
                          semrush_rank} ... ] sorted by ym,
             "latest_ym": str, "domain": str}
    """
    domain = config["brand_name_map"][config["brand"]]["web"]
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            return {"months": [], "latest_ym": None, "domain": domain}
        hmap = {str(h).strip().lower(): i for i, h in enumerate(header)}

        def idx(*names):
            for n in names:
                if n in hmap:
                    return hmap[n]
            return None

        i_month = idx("month")
        i_domain = idx("domain", "target")
        i_as = idx("authority score")
        i_rank = idx("semrush rank")
        i_traf = idx("organic traffic")
        i_kw = idx("organic keywords")
        i_bl = idx("backlinks")
        i_rd = idx("referring domains", "ref domains")
        if i_month is None or i_domain is None:
            raise ValueError("SEO export missing Month/Domain columns: %s" % path)

        def cell(row, i):
            return row[i] if i is not None and i < len(row) else None

        by_ym = {}
        for row in reader:
            if not row or i_domain >= len(row):
                continue
            dom = (row[i_domain] or "").strip().lower()
            if dom != domain.lower():
                continue
            ym = _month_label_to_ym(cell(row, i_month))
            if ym is None:
                continue
            by_ym[ym] = {
                "ym": ym,
                "month_label": str(cell(row, i_month)).strip(),
                "organic_traffic": _to_int(_clean_number(cell(row, i_traf))),
                "organic_keywords": _to_int(_clean_number(cell(row, i_kw))),
                "backlinks": _to_int(_clean_number(cell(row, i_bl))),
                "referring_domains": _to_int(_clean_number(cell(row, i_rd))),
                "authority_score": _to_int(_clean_number(cell(row, i_as))),
                "semrush_rank": _to_int(_clean_number(cell(row, i_rank))),
            }

    months = [by_ym[k] for k in sorted(by_ym)]
    return {"months": months,
            "latest_ym": months[-1]["ym"] if months else None,
            "domain": domain}


# --------------------------------------------------------------------------
# AEO (Degreed-only tab + Overall SOV channel) -- section 13b / 14
# --------------------------------------------------------------------------
def parse_aeo(path):
    """
    Read the AI Visibility grid (SEMrush AI Visibility + Microsoft Clarity SoA).

    Layout: a 'Year' row and a 'Month' row across the columns, then one row per
    metric. We read TOTALS ONLY (section 13b) -- the indented per-engine
    sub-rows (Chat GPT, AI Overview, AI Mode, Gemini) and the Degreed /
    competitor mention sub-rows beneath Share of Authority are ignored.

    Returns {"months": [ {ym, year, month_name, score, mentions, citations,
                          cited_pages, share_of_authority} ... ] sorted by ym,
             "soa_by_ym": {ym: fraction},   # Share of Authority as a fraction
             "latest_ym": str}
    """
    rows = []
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    year_row = month_row = None
    for r in rows:
        if not r:
            continue
        head = (r[0] or "").strip().lower()
        if head == "year":
            year_row = r
        elif head == "month":
            month_row = r
        if year_row and month_row:
            break
    if month_row is None:
        return {"months": [], "soa_by_ym": {}, "latest_ym": None}

    # Column -> ym. Forward-fill the year across blank cells.
    col_ym = {}
    cur_year = None
    for c in range(1, len(month_row)):
        if year_row and c < len(year_row):
            yr = (year_row[c] or "").strip()
            if re.match(r"^\d{4}$", yr):
                cur_year = int(yr)
        name = (month_row[c] or "").strip().lower()
        num = _MON_FULL.get(name) or _MON_ABBR.get(name[:3]) if name else None
        if num and cur_year:
            col_ym[c] = "%04d-%02d" % (cur_year, num)

    # Metric label (top-level, no leading whitespace) -> internal key.
    metric_key = {
        "total score": "score",
        "ai visibility score": "score",
        "cited pages": "cited_pages",
        "citations": "citations",
        "mentions": "mentions",
        "share of authority": "share_of_authority",
    }

    data = {ym: {} for ym in col_ym.values()}
    for r in rows:
        if not r or r[0] is None:
            continue
        raw_label = r[0]
        if raw_label != raw_label.lstrip():
            continue  # indented sub-row (per-engine / mention split) -> skip
        key = metric_key.get(raw_label.strip().lower())
        if key is None:
            continue
        for c, ym in col_ym.items():
            val = r[c].strip() if c < len(r) and r[c] is not None else ""
            if val == "":
                continue
            if key == "share_of_authority":
                num = _clean_number(val)  # strips the % sign
                data[ym][key] = (num / 100.0) if num is not None else None
            else:
                data[ym][key] = _to_int(_clean_number(val))

    months = []
    soa_by_ym = {}
    for ym in sorted(data):
        d = data[ym]
        if not any(v is not None for v in d.values()):
            continue  # future / empty month
        y, mo = ym.split("-")
        months.append({
            "ym": ym,
            "year": int(y),
            "month_name": _MON_NAME[int(mo)] + " " + y,
            "score": d.get("score"),
            "mentions": d.get("mentions"),
            "citations": d.get("citations"),
            "cited_pages": d.get("cited_pages"),
            "share_of_authority": d.get("share_of_authority"),
        })
        if d.get("share_of_authority") is not None:
            soa_by_ym[ym] = d["share_of_authority"]

    return {"months": months, "soa_by_ym": soa_by_ym,
            "latest_ym": months[-1]["ym"] if months else None}


# --------------------------------------------------------------------------
# Lead Capture (Degreed-only tab) -- HubSpot Form Performance, section 15
# --------------------------------------------------------------------------
def _daterange_to_ym(text):
    """'Date range: 06/01/26 - 06/30/26' -> ('2026-06', '06/01/26 - 06/30/26')."""
    m = re.search(r"Date range:\s*(\d{2})/(\d{2})/(\d{2})\s*-\s*(.+)$", text)
    if not m:
        return None, None
    mo, _dd, yy = int(m.group(1)), m.group(2), int(m.group(3))
    ym = "%04d-%02d" % (2000 + yy, mo)
    full = "%s/%s/%s - %s" % (m.group(1), _dd, m.group(3), m.group(4).strip())
    return ym, full


def parse_hubspot_forms(paths, config):
    """
    Read every monthly HubSpot Form Performance CSV (section 15a) and compute
    the monthly lead-capture rate (section 15b):

        rate = (sum of Submissions / sum of Form Views) * 100

    across all COUNTED forms. Two rates are recorded per month: the headline
    rate with the exclusion filter applied, and the raw all-forms rate as a
    check. The per-form 'Submissions Per Form View' column is never averaged.

    Returns {"months": [ {ym, month_label, date_range, filtered_rate,
                          raw_rate, filtered_submissions, filtered_views,
                          raw_submissions, raw_views, excluded_forms,
                          counted_forms, top_forms, source_file} ...] by ym,
             "patterns": [...]}
    """
    patterns = [p.lower() for p in config.get("exclude_form_name_patterns", [])]

    def excluded(name):
        low = name.lower()
        return any(p in low for p in patterns)

    months = []
    for path in paths:
        if not os.path.exists(path):
            continue
        with open(path, "r", newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header is None:
                continue

            date_range = None
            ym = None
            raw_sub = raw_views = filt_sub = filt_views = 0
            excl_count = counted = 0
            top = []
            for row in reader:
                if not row:
                    continue
                first = (row[0] or "").strip()
                if first.startswith("Date range:"):
                    ym, date_range = _daterange_to_ym(first)
                    continue
                if first.startswith("Generated on") or first == "":
                    continue
                if len(row) < 4:
                    continue
                name = (row[1] or "").strip()
                sub = _to_int(row[2])
                views = _to_int(row[3])
                if sub is None or views is None:
                    continue
                raw_sub += sub
                raw_views += views
                if excluded(name):
                    excl_count += 1
                    continue
                counted += 1
                filt_sub += sub
                filt_views += views
                if sub > 0:
                    top.append({"name": name, "submissions": sub, "views": views})

        if ym is None:
            # fall back to the filename month if the trailer row was absent
            base = os.path.basename(path)
            mm = re.search(r"_([A-Za-z]{3})(\d{2})", base)
            if mm:
                num = _MON_ABBR.get(mm.group(1).lower())
                if num:
                    ym = "%04d-%02d" % (2000 + int(mm.group(2)), num)

        top.sort(key=lambda x: x["submissions"], reverse=True)
        y, mo = (ym.split("-") if ym else (None, None))
        months.append({
            "ym": ym,
            "month_label": (_MON_NAME[int(mo)] + " " + y) if ym else os.path.basename(path),
            "date_range": date_range,
            "filtered_rate": (filt_sub / filt_views * 100) if filt_views else None,
            "raw_rate": (raw_sub / raw_views * 100) if raw_views else None,
            "filtered_submissions": filt_sub,
            "filtered_views": filt_views,
            "raw_submissions": raw_sub,
            "raw_views": raw_views,
            "excluded_forms": excl_count,
            "counted_forms": counted,
            "top_forms": top[:8],
            "source_file": os.path.basename(path),
        })

    months.sort(key=lambda m: (m["ym"] or ""))
    return {"months": months, "patterns": config.get("exclude_form_name_patterns", [])}
